# -*- coding: utf-8 -*-


# 2min --> 1 000 000 de comptes



from openpyxl import Workbook, load_workbook
from win10toast import ToastNotifier
import random
import codecs

wb = Workbook()
ws = wb.active
ws.title = "Accounts"
#print(wb.sheetnames)

n = ToastNotifier()
nb = int(input("Combien de comptes ? "))
#file = open("accounts.txt", "a+")

names_file = open("account_generator-main\\noms.txt", "r", encoding='utf8')
lines = names_file.readlines()
def nom():
    liste = []
    for line in lines:
        liste.append(line.strip())
    names_file.close
    return random.choice(liste)

firsts_name_file = open("account_generator-main\\prenoms.txt", "r", encoding='utf8')
lines = firsts_name_file.readlines()
def prenom():
    liste = []
    for line in lines:
        liste.append(line.strip())
    firsts_name_file.close
    return random.choice(liste)

def birth_date():
    day = random.randint(1, 30)
    month = random.randint(1, 12)
    year = random.randint(1950, 2002)
    date = f"{day}/{month}/{year}"
    return date

def mdp():
    return

#file.write("Mail        Nom     Prenom      Date de naissance\n")
ws.append(["Mail", "Nom", "Prenom", "Date de naissance", "Mot de passe"])
print()
min = 0
for i in range(nb):
    prenom1 = prenom()
    nom1 = nom()
    date = birth_date()
    mail = prenom1 + "." + nom1 +"@gmail.com"
    #file.write(prenom1 + "." + nom1 +"@gmail.com   " + nom1 + "    " + prenom1 + "  " + date + "\n")
    ws.append([mail, nom1, prenom1, date])
    pourcent = int((i/nb)*100)
    if pourcent > min:
        print(pourcent, "%", end="\r")
        min = pourcent
#file.close


wb.save("accounts.xlsx")
print("100 %")
print(f"Génération de {nb} comptes terminée...")
n.show_toast("Account_generator", "La génération des comptes est terminée", duration = 10)


